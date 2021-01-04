from openpyxl import load_workbook, Workbook
from random import *

ck = Workbook(write_only=True)
# excel çalıma kitabı oluşturuldu


class sinif():
    def __init__(self):
        self.sinif_ogrencileri = []
    def sinif_tanimla(self,sinif_adi):
        self.sinif_adi = sinif_adi
    def ogr_sayi(self,ogr_sayisi):
        self.ogr_sayisi = ogr_sayisi
    def ogrenci_sayisi(self):
        return len(self.sinif_ogrencileri)
    def ogrenci_ekle(self,ogrenci):
        self.sinif_ogrencileri.append(ogrenci)
    def gidilecek_sinif_ekle(self,ogr_no):
        self.gidilecek_sinif=ogr_no
    def gsl(self):
        global ck
        eck = ck.create_sheet(self.sinif_adi)
        for ogrenci in self.sinif_ogrencileri:
            ogrenci_bilgileri = "öğrenci no:{}, öğrenci adı : {}, öğrenci soyadı : {}, öğrencinin sinifi : {}, Öğrencinin Gideceği sinif : {}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3],ogrenci[4])
            # print(ogrenci)
            eck.append(ogrenci_bilgileri)
    def ogrenci_listesi(self):
        for ogrenci in self.sinif_ogrencileri:
            pass
            # print("öğrenci no:{:10}, öğrenci adı : {:20}, öğrenci soyadı : {:20}, öğrencinin sinifi : {:10}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3]))
        return self.sinif_ogrencileri

class f_sinif():
    global ck
    def __init__(self):
        self.sinif_ogrencileri = []
    def sinif_tanimla(self,sinif_adi):
        self.sinif_adi = sinif_adi
        if not(self.sinif_adi in ck.sheetnames):
            ck.create_sheet(self.sinif_adi)
    def ogr_sayi(self,ogr_sayisi):
        self.ogr_sayisi = len(self.sinif_ogrencileri)
    def ogrenci_ekle(self,ogrenci):
        self.sinif_ogrencileri.append(ogrenci)

    def gidilecek_sinif_ekle(self,ogr_no):
        self.gidilecek_sinif=ogr_no
    def gsl(self):
        for ogrenci in self.sinif_ogrencileri:
            pass
            # print("öğrenci no:{}, öğrenci adı : {}, öğrenci soyadı : {}, öğrencinin sinifi : {}, Öğrencinin Gideceği sinif : {}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3],ogrenci[4]))
    def ogrenci_listesi(self):

        # for k in range(2):
        #     ck[self.sinif_adi].append([" "])
        for ogrenci in self.sinif_ogrencileri:
            # print("no:{}, adı : {:3}, soyadı : {}, sinifi : {}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3]))
            bilgiler = "no:{}, adı : {:3}, soyadı : {}, sinifi : {}".format(ogrenci[0],ogrenci[1],ogrenci[2],ogrenci[3])
            ck[self.sinif_adi].append(ogrenci)
        return self.sinif_ogrencileri
    def sos(self):
        dokuzlar = 0
        onlar = 0
        onbirler = 0
        for ogrenci in self.sinif_ogrencileri:
            if "9" in ogrenci[3]:
                dokuzlar +=1
            elif "10" in ogrenci[3]:
                onlar +=1
            elif "11" in ogrenci[3]:
                onbirler +=1
        for k in range(2):
            ck[self.sinif_adi].append([" "])
        icerik = "9 Siniflar = {},  10 Siniflar = {},   11 Siniflar = {}".format(dokuzlar,onlar,onbirler)
        ck[self.sinif_adi].append(icerik.split(","))
        return ck[self.sinif_adi]
        # return print("9 Siniflar = {},  10 Siniflar = {},   11 Siniflar = {},  ".format(dokuzlar,onlar,onbirler))

wb = load_workbook(filename='okullistesi.xlsx', read_only=True)
ws = wb['Sheet1']
# print(ws.max_row)
ogrenciler = []
sinif_sayilari = []

for deger in range(2,ws.max_row):
    hdegeri = "P"+str(deger-1)
    if ws[hdegeri].value != None:
        smevcut = ws[hdegeri].value
        sinif_sayilari.append(smevcut)
baslangic = 2
siniflar = ["sinif_9a","sinif_9b","sinif_9c",
            "sinif_10a","sinif_10b","sinif_10c",
            "sinif_11a","sinif_11b","sinif_11c","sinif_11d"]

# siniflar = ["sinif_9a","sinif_9b","sinif_9c",
#             "sinif_10a","sinif_10b","sinif_10c",
#             "sinif_11a","sinif_11b","sinif_11c","sinif_11d",
#             "sinif_12a","sinif_12b","sinif_12c","sinif_12d","sinif_12e"]

ssiniflar = iter(siniflar)
fsinif = {}
for ogrenci in siniflar:
    exec(ogrenci + " = sinif()")
    exec("f_" + ogrenci + " = f_sinif()")
# print(sum(sinif_sayilari))
for ilerleme in sinif_sayilari:
    x_sinif = str(next(ssiniflar))
    fsinif[x_sinif] = ilerleme
    o_sinif = eval(x_sinif)
    o_sinif.sinif_tanimla(x_sinif)
    o_sinif.ogr_sayi(ilerleme)
    # print(baslangic,ilerleme)
    for deger in range(baslangic, baslangic+ilerleme, 1):
        ssatir = ws[deger]
        ogrenci = []
        for cell in ssatir[2], ssatir[5], ssatir[9]:
            ogrenci.append(cell.value)
        ogrenci.append(x_sinif)
        o_sinif.ogrenci_ekle(ogrenci)
        # print(ogrenci)

    baslangic = baslangic + ilerleme + 3
fsiniflar = iter(siniflar)
for ilerleme in sinif_sayilari:
    x_sinif = "f_"+str(next(fsiniflar))
    # fsinif[x_sinif]=ilerleme
    o_sinif = eval(x_sinif)
    o_sinif.sinif_tanimla(x_sinif)
    o_sinif.ogr_sayi(ilerleme)

# esinif eski ysinif yeni sinif
def karisik_listele(esinif,ysiniff):
    esinif = eval(esinif)
    ysinif = eval("f_"+ysiniff)
    sayi = esinif.ogrenci_sayisi()
    rast = randrange(sayi)
    if len(ysinif.sinif_ogrencileri) < fsinif[ysiniff]:

        global ck
        sinif_adi = esinif.sinif_adi
        if not(sinif_adi in ck.sheetnames):
            eck = ck.create_sheet(sinif_adi)

        ysinif.ogrenci_ekle(esinif.sinif_ogrencileri[rast])
        # print(esinif.sinif_ogrencileri[rast]," = >",ysiniff)
        esinif.sinif_ogrencileri[rast].append(" = >" + ysiniff)
        ck[sinif_adi].append(esinif.sinif_ogrencileri[rast])
        esinif.sinif_ogrencileri.pop(rast)
        global toplam
        toplam += 1
        return True
    else:
        return False

# print(fsinif)
bsiralama = set(sinif_sayilari)
# bsiralama = sorted(sinif_sayilari)
# bsiralama = bsiralama[::-1]
ylistele = {}
yslist=[]
for xmen in range(len(bsiralama)):
    l = 0
    for ymen in bsiralama:
        for ara in fsinif:
            if (not (ara in ylistele)) and (ymen == fsinif[ara]):
                ylistele[ara] = ymen
                yslist.append(ara)
                # bsiralama.pop(l)
        # l = l+1

# print(ylistele)
z = fsinif
toplam = 0
# print(fsinif)





while toplam < sum(sinif_sayilari):
    kontrol = 0
    for a in yslist:
        # print(a, ylistele[a])
        o_sinif = eval(a)

        mevcut = 1

        while len(o_sinif.sinif_ogrencileri) > 0:
            # print(mevcut, end=" ")
            k = yslist[(kontrol%10)]
            b = k
            if karisik_listele(a, b):
                mevcut = mevcut + 1
            kontrol += 1

# print("devam")

ert = 0
for a in z:
    # print(a,"=>",z[a])
    y_sinif = eval("f_"+a)
    # print(y_sinif.sinif_adi, len(y_sinif.sinif_ogrencileri))
    # ert +=len(y_sinif.ogrenci_listesi())
    y_sinif.ogrenci_listesi()
    y_sinif.sos()
#     print("\n"*2)
# print(ert)



ck.save('new_big_file.xlsx')
