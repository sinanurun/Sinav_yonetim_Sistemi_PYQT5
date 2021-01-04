# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'subebilgileriguncelleme.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook, Workbook
from random import *

import sys
from db_islemleri import * #db işlemleri sayfasını aldık


class Sube_bilgi_guncelle(QtWidgets.QWidget):
    def __init__(self):
        super(Sube_bilgi_guncelle,self).__init__()

        self.setObjectName("Form")
        self.resize(623, 355)

        self.query = session.query(SubeBilgileri)
        self.tablo = QtWidgets.QTableWidget(self)
        self.tablo.setGeometry(QtCore.QRect(20,30,600,250))
        self.tablo.setRowCount(self.query.count())
        self.tablo.setColumnCount(3)
        self.tablo.setHorizontalHeaderLabels(["Şube Adı Eski","Şube Mevcutları","Şube Öğrencileri"])
        self.tablo.setColumnWidth(0, 100)
        self.tablo.setColumnWidth(1, 100)
        self.tablo.setColumnWidth(2, 400)



        for dd in range(self.query.count()):
            icerik = self.query.filter(SubeBilgileri.sube_id == dd + 1).first()
            self.tablo.setItem(dd,0,QtWidgets.QTableWidgetItem(icerik.sube_adi))
            self.tablo.setItem(dd, 1, QtWidgets.QTableWidgetItem(str(icerik.sube_mevcut)))
            self.tablo.setItem(dd, 2, QtWidgets.QTableWidgetItem(icerik.sube_ogrenci))

        self.guncelle = QtWidgets.QPushButton(self)
        self.guncelle.setGeometry(QtCore.QRect(530, 320, 75, 23))
        self.guncelle.setObjectName("guncelle")
        self.renciListesiEkLabel = QtWidgets.QLabel(self)
        self.renciListesiEkLabel.setGeometry(QtCore.QRect(200, 280, 275, 22))
        self.renciListesiEkLabel.setObjectName("renciListesiEkLabel")
        self.listeal = QtWidgets.QPushButton(self)
        self.listeal.setGeometry(QtCore.QRect(500, 280, 110, 23))
        self.listeal.setObjectName("listeal")

        self.retranslateUi()
        QtCore.QMetaObject.connectSlotsByName(self)

        # çalıştıılacak fonksiyonlar

        # self.tablo.cellChanged.connect(self.degisti)
        self.listeal.clicked.connect(self.listealfonk)

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("Form", "Şube bilgileri Güncelleme"))
        self.guncelle.setText(_translate("Form", "Güncelle"))
        self.renciListesiEkLabel.setText(_translate("Form", "Şube Mevcutları ve Öğrenci Lislerini Excelden Çek ==>>"))
        self.listeal.setText(_translate("Form", "Listeyi Al"))

    @QtCore.pyqtSlot()
    def degisti(self):
        try:
            gelen = self.sender().currentRow()
            print(gelen)

        except:
            print("hata")


    def listealfonk(self):



        wb = load_workbook(filename='okullistesi.xlsx', read_only=True)
        ws = wb['Sheet1']
        # print(ws.max_row)
        ogrenciler = []
        sinif_sayilari = []

        for deger in range(2, ws.max_row):
            hdegeri = "P" + str(deger - 1)
            if ws[hdegeri].value != None:
                smevcut = ws[hdegeri].value
                sinif_sayilari.append(smevcut)
        baslangic = 2
        siniflar=[]
        for dd, a in zip(range(self.query.count()),sinif_sayilari):

            siniflar.append(self.tablo.item(dd,0).text())
            self.tablo.setItem(dd,1,QtWidgets.QTableWidgetItem(str(a)))

        ssiniflar = iter(siniflar)

        # print(sum(sinif_sayilari))
        session.query(OgrenciListeleri).delete(synchronize_session='evaluate')
        session.commit()
        session.query(SubeBilgileri).delete(synchronize_session='evaluate')
        session.commit()

        for ilerleme,satir in zip(sinif_sayilari,range(len(sinif_sayilari))):
            ogrenci = []
            for deger in range(baslangic, baslangic + ilerleme, 1):
                ssatir = ws[deger]
                ogr = []
                for cell in ssatir[2], ssatir[5], ssatir[9]:
                    ogr.append(cell.value)
                ogrenci.append(ogr)
                ogren = OgrenciListeleri(ogrenci_adi=ogr[1]+" "+ogr[2],ogrenci_no=int(ogr[0]),sube_id=satir+1)
                session.add(ogren)
                session.commit()
            self.tablo.setItem(satir, 2, QtWidgets.QTableWidgetItem(str(ogr)))
            subeler = []

            sube = SubeBilgileri(sube_adi=self.tablo.item(satir,0).text(),sube_mevcut=ilerleme,sube_ogrenci=str(ogrenci))
            session.add(sube)
            session.commit()
            baslangic = baslangic + ilerleme + 3

# if __name__ == "__main__":
#     import sys
#     app = QtWidgets.QApplication(sys.argv)
#     Form = QtWidgets.QWidget()
#     ui = Sube_bilgi_guncelle()
#     ui.show()
#     sys.exit(app.exec_())

