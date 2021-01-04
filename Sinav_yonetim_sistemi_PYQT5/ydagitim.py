# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'yenidagitim.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from ydagitim_ui import *
import sys
from db_islemleri import * #db işlemleri sayfasını aldık
from PyQt5.QtWidgets import QComboBox,QMessageBox,QLabel
from openpyxl import load_workbook, Workbook
from random import *
from datetime import date
# from fpdf import FPDF




class Yeni_Dagitim(QtWidgets.QWidget,Ui_Form):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.setupUi(self)

        # self.sinav_adi.textChanged.connect(self.deneme)
        self.sinav_sayisi.valueChanged.connect(self.tablo_olustur01)
        self.sinav_tablosu_01.setRowCount(1)
        self.sinav_tablosu_01.setItem(0, 0, QtWidgets.QTableWidgetItem(str(1) + ". Sınav Adı :"))
        self.sinav_tablosu_01.setItem(0, 1, QtWidgets.QTableWidgetItem("Sınav Adı Giriniz 01"))

        # self.sube_bilgileri.connect(self.tablo_olustur02)
        self.tab_bilgialanlari.currentChanged.connect(self.secim)
        self.dagitim.clicked.connect(self.dagit)

    def secim(self):
        self.alan = self.sender()
        self.secili_alan = self.alan.currentIndex()
        if self.secili_alan == 0:
            pass
        elif self.secili_alan == 1 :
            return self.tablo_olustur02()
        elif self.secili_alan == 2:
            return self.tablo_olustur03()


    def tablo_olustur01(self):
        self.gelen = self.sender()
        self.sinav_tablosu_01.setRowCount(int(self.gelen.text()))
        for dd in range(int(self.gelen.text())):
           self.sinav_tablosu_01.setItem(dd, 0, QtWidgets.QTableWidgetItem(str(dd+1)+". Sınav Adı :"))
           self.sinav_tablosu_01.setItem(dd, 1, QtWidgets.QTableWidgetItem("Sınav Adı Giriniz "+str(dd+1)))

    def tablo_olustur02(self):

        self.sinav_tablosu_02.setColumnCount(2)
        baslik=[self.sinav_tablosu_02.horizontalHeaderItem(0).text(),"Sınav Seçimi"]
        self.ders=[]
        self.sinav_subeleri = {}
        for dd in range(self.sinav_sayisi.value()):
            self.ders.append(self.sinav_tablosu_01.item(dd,1).text())
        self.sinav_tablosu_02.setHorizontalHeaderLabels(baslik)

        self.sinav_tablosu_02.setRowCount(session.query(SubeBilgileri).count())

        self.siniflar = session.query(SubeBilgileri).all()
        self.sayilar = session.query(SubeBilgileri).count()
        for dd , sinif in zip(range(self.sayilar),self.siniflar):
            self.sinav_tablosu_02.setItem(dd,0,QtWidgets.QTableWidgetItem(sinif.sube_adi))
            combo = QComboBox(self)
            combo.setObjectName(str(dd))
            combo.addItem("Sınav Seçiniz")
            combo.addItems(self.ders)
            self.sinav_tablosu_02.setCellWidget(dd,1,combo)
            self.sinav_subeleri[dd]=[sinif.sube_id,combo,sinif.sube_adi]
            # for kk in ders[1:]:
            #
            #     kutu = QtWidgets.QRadioButton(self)
            #     kutu.setObjectName(str(dd))
            #     kutu.setText(kk)
            #
            #
            #     kutu.toggled.connect(self.yazdir)
            #     x+=1
            # self.sinav_tablosu_02.setItem(dd, 0, QtWidgets.QTableWidgetItem("ddfsfsf"))
        # print(self.ders)

    def tablo_olustur03(self):

        self.fsiniflar = session.query(FizikiSinif).all()
        self.fsayilar = session.query(FizikiSinif).count()
        self.sinav_tablosu_03.setRowCount(self.fsayilar)


        self.fiziki_siniflar = {}
        for sayi , sinif in zip(range(self.fsayilar),self.fsiniflar):
            kutu = QtWidgets.QCheckBox(self)
            kutu.setObjectName(str(sayi))
            self.fiziki_siniflar[sayi]=[sinif.fsinif_id,kutu,sinif.fsinif_adi]
            self.sinav_tablosu_03.setItem(sayi,0,QtWidgets.QTableWidgetItem(sinif.fsinif_adi))
            self.sinav_tablosu_03.setCellWidget(sayi, 1, kutu)
        #
        # print("sınav yerleri")

    def dagit(self):

        dialog = QMessageBox(self)
        islem = QLabel(dialog, text="İşlem Bitene Kadar Bekleyiniz")
        dialog.show()

        # print(self.sinav_adi.text())

        # sınav olacak sınıflar ve olacakları sınavlar

        sinav_ogrencileri = {}
        a = 1
        for dd in range(self.sayilar):
            if self.sinav_subeleri[dd][1].currentIndex() !=0:
                # print(self.sinav_subeleri[dd][0],self.sinav_subeleri[dd][1].currentText(),self.sinav_subeleri[dd][1].currentIndex())
                # sinav_subeleri[self.sinav_subeleri[dd][0]] = self.sinav_subeleri[dd][1].currentText()
                sinav_subesi = session.query(OgrenciListeleri).filter(OgrenciListeleri.sube_id==self.sinav_subeleri[dd][0]).all()
                shuffle(sinav_subesi)
                for k in sinav_subesi:
                    sinav_ogrencileri[a]=[k.ogrenci_id,k.ogrenci_adi,k.sube_id,
                                                     self.sinav_subeleri[dd][1].currentText(),self.sinav_subeleri[dd][1].currentIndex()]
                    a+=1
        # print(len(sinav_ogrencileri))
        # print(sinav_ogrencileri)
        sinav_mekanlari = {}
        #sınav olacak olan fiziki sınıflar
        for dd in range(self.fsayilar):
            if self.fiziki_siniflar[dd][1].isChecked():
                # print(self.fiziki_siniflar[dd][1].objectName(),self.fiziki_siniflar[dd][0])
                mevcut = session.query(FizikiSinif).filter(FizikiSinif.fsinif_id==int(self.fiziki_siniflar[dd][0])).first()
                mevcut = mevcut.fsinif__mevcut
                sinav_mekanlari[self.fiziki_siniflar[dd][0]]=mevcut

        # print(sum(sinav_mekanlari.values()))
        # print(sinav_mekanlari)
        if len(sinav_ogrencileri)>sum(sinav_mekanlari.values()):
            dialog = QMessageBox(self)
            dialog.setWindowTitle("Atama Hatası")
            dialog.setText("Fiziki Mekan Yetersiz")
            dialog.show()

        else:

            l = 1
            for a in range(1,len(self.ders)+1):
                key = list(sinav_mekanlari.keys())
                z = 0

                for k in sinav_ogrencileri:
                    c = key[z]
                    while sinav_ogrencileri[k][4] == a and len(sinav_ogrencileri[k])==5:
                        if sinav_mekanlari[c]>0:
                            sinav_ogrencileri[k].append(c)
                            sinav_mekanlari[c] = sinav_mekanlari[c]-1
                            # print(z,l)
                            # print(sinav_ogrencileri[k])
                            l +=1
                            z += 1
                            z = z % len(sinav_mekanlari)
                        else:
                           z += 1
                           z = z % len(sinav_mekanlari)
                           c = key[z]

        # sube için excel
        sl = Workbook(write_only=True)
        fl = Workbook(write_only=True)
        tl = Workbook(write_only=True)
        # pdf1 = FPDF()
        # pdf1.set_font("Arial", size=12)
        # pdf2= FPDF()
        # pdf2.set_font("Arial", size=12)
        a,b =0,0
        for ogrenci in sinav_ogrencileri:
            ogrenci_id = sinav_ogrencileri[ogrenci][0]
            ogrenci_adi = sinav_ogrencileri[ogrenci][1]
            ogrenci_sube = sinav_ogrencileri[ogrenci][2]
            sinav_adi = sinav_ogrencileri[ogrenci][3]
            sinavyeri_id = sinav_ogrencileri[ogrenci][5]

            if not(self.sinav_subeleri[ogrenci_sube-1][2] in sl.sheetnames):
                sl.create_sheet(self.sinav_subeleri[ogrenci_sube-1][2])
                tl.create_sheet(self.sinav_subeleri[ogrenci_sube - 1][2])
                sl[self.sinav_subeleri[ogrenci_sube - 1][2]].append(["Öğrenci Adi","Şubesi","Sinav Adi","Sınav Yeri"])
                tl[self.sinav_subeleri[ogrenci_sube - 1][2]].append(
                    ["Öğrenci Adi", "Şubesi", "Sinav Adi", "Sınav Yeri"])

                # pdf1.add_page()
                #
                # col_width1 = pdf1.w / 4.5
                # row_height1 = pdf1.font_size
                # pdf1.cell(col_width1, row_height1 * 1, txt="{:30}-{:10}-{:10}-{:10}".format("Öğrenci Adi","Şubesi","Sinav Adi","Sınav Yeri"), border=1)

            sl[self.sinav_subeleri[ogrenci_sube-1][2]].append((ogrenci_adi,self.sinav_subeleri[ogrenci_sube - 1][2],
                                                               sinav_adi,self.fiziki_siniflar[sinavyeri_id-1][2]))
            # pdf1.cell(col_width1, row_height1 * 1,
            #           txt="{:30}-{:10}-{:10}-{:10}".format(ogrenci_adi,self.sinav_subeleri[ogrenci_sube - 1][2],
            #                                                    sinav_adi,self.fiziki_siniflar[sinavyeri_id-1][2]),
            #           border=1)
            #
            tl[self.sinav_subeleri[ogrenci_sube-1][2]].append((ogrenci_adi,self.sinav_subeleri[ogrenci_sube - 1][2],
                                                               sinav_adi,self.fiziki_siniflar[sinavyeri_id-1][2]))


            if not(self.fiziki_siniflar[sinavyeri_id-1][2] in fl.sheetnames):
                fl.create_sheet(self.fiziki_siniflar[sinavyeri_id-1][2])
                tl.create_sheet(self.fiziki_siniflar[sinavyeri_id-1][2])
                fl[self.fiziki_siniflar[sinavyeri_id-1][2]].append(
                    ["Öğrenci Adi", "Şubesi", "Sinav Adi", "Sınav Yeri"])
                tl[self.fiziki_siniflar[sinavyeri_id-1][2]].append(
                    ["Öğrenci Adi", "Şubesi", "Sinav Adi", "Sınav Yeri"])
                # pdf2.add_page()
                #
                # col_width1 = pdf2.w / 4.5
                # row_height1 = pdf2.font_size

            fl[self.fiziki_siniflar[sinavyeri_id-1][2]].append((ogrenci_adi, self.sinav_subeleri[ogrenci_sube - 1][2],

                                                                 sinav_adi, self.fiziki_siniflar[sinavyeri_id - 1][2]))
            tl[self.fiziki_siniflar[sinavyeri_id-1][2]].append((ogrenci_adi, self.sinav_subeleri[ogrenci_sube - 1][2],

                                                                 sinav_adi, self.fiziki_siniflar[sinavyeri_id - 1][2]))


        sl.save(self.sinav_adi.text() + '_ Şube Listeleri.xlsx')
        fl.save(self.sinav_adi.text() + '_ Sinav Yeri Listeleri.xlsx')
        tl.save(self.sinav_adi.text() + '_ Toplu Listeleri.xlsx')
        # pdf1.output("multipage_simple.pdf")
        # print(sinav_ogrencileri)
        # print(sinav_mekanlari)

        kayit = Dagitim(dagitim_adi=self.sinav_adi.text()+" "+str(date.today()),
                        dagitim_tarihi=str(date.today()),
                        dsinif_listesi=str(self.sinav_adi.text() + '_ Şube Listeleri.xlsx'),
                        dfsinif_listesi=str(self.sinav_adi.text() + '_ Sinav Yeri Listeleri.xlsx'),
                        dtoplu_liste=str(self.sinav_adi.text() + '_ Toplu Listeleri.xlsx')
                        )

        session.add(kayit)
        session.commit()
        try:
            dialog.close()
        except:
            pass
        finally:
            dialog = QMessageBox(self)
            islem = QLabel(dialog, text="İşlem Başarılı")
            dialog.buttonClicked.connect(lambda :self.parent().baslat())
            dialog.show()


if __name__=="__main__":
    app = QtWidgets.QApplication(sys.argv)
    giris = Yeni_Dagitim()

    giris.show()
    sys.exit(app.exec_())